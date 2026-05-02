import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime


def export_stats_to_xlsx(users: list[dict]) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Статистика"

    hdr_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill("solid", fgColor="4472C4")
    hdr_align = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side("thin"), right=Side("thin"),
        top=Side("thin"), bottom=Side("thin"),
    )

    headers = ["#", "User ID", "Username", "Имя", "Вступивших", "Ссылка"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font, c.fill, c.alignment, c.border = hdr_font, hdr_fill, hdr_align, border

    for idx, u in enumerate(users, 1):
        row = idx + 1
        ws.cell(row=row, column=1, value=idx).border = border
        ws.cell(row=row, column=2, value=u["user_id"]).border = border
        ws.cell(row=row, column=3, value=u.get("username") or "—").border = border
        ws.cell(row=row, column=4, value=u.get("first_name") or "—").border = border

        cnt = ws.cell(row=row, column=5, value=u.get("referral_count", 0))
        cnt.border = border
        cnt.alignment = Alignment(horizontal="center")
        if u.get("referral_count", 0) > 0:
            cnt.font = Font(bold=True, color="2E7D32")

        ws.cell(row=row, column=6, value=u.get("invite_link") or "—").border = border

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 40
    ws.auto_filter.ref = f"A1:F{len(users) + 1}"

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    path = os.path.join(os.path.dirname(__file__), f"stats_{ts}.xlsx")
    wb.save(path)
    return path